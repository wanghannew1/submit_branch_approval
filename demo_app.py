import os
import re
import json
import time
import logging
from datetime import datetime
from io import BytesIO
from pathlib import Path
import uuid

import requests
import openpyxl
import streamlit as st
from dotenv import load_dotenv

import lark_oapi as lark
from lark_oapi.api.contact.v3 import BatchGetIdUserRequest, BatchGetIdUserRequestBody
from lark_oapi.api.contact.v3 import BasicBatchUserRequest, BasicBatchUserRequestBody
from lark_oapi.api.approval.v4 import CreateInstanceRequest, InstanceCreate
from requests_toolbelt import MultipartEncoder

# ── Logging Setup ───────────────────────────────────────────────
LOG_DIR = Path(__file__).parent / "log"
LOG_DIR.mkdir(exist_ok=True)

logger = logging.getLogger("submit_approval_feishu")
logger.setLevel(logging.DEBUG)

# File handler: all debug messages go to a daily rotating file
log_file = LOG_DIR / f"app_{datetime.now().strftime('%Y%m%d')}.log"
fh = logging.FileHandler(log_file, encoding="utf-8")
fh.setLevel(logging.DEBUG)

# Console handler: info and above
ch = logging.StreamHandler()
ch.setLevel(logging.INFO)

formatter = logging.Formatter(
    "%(asctime)s [%(levelname)s] %(message)s", datefmt="%Y-%m-%d %H:%M:%S"
)
fh.setFormatter(formatter)
ch.setFormatter(formatter)
logger.addHandler(fh)
logger.addHandler(ch)

# ── Environment ─────────────────────────────────────────────────
load_dotenv()

DEFAULT_CONFIG = {
    "excel": {
        "title_row": 1,           # 报表名所在行（1-based）
        "unit_name_row": 2,       # 单位名所在行（1-based）；0 表示无此行，单位名从标题行用正则提取
        "header_start_row": 3,    # 表头起始行（1-based）
        "header_row_count": 3,    # 表头占几行（适配多级合并表头）
        "summary_row_marker": "合计",
        "unit_name_patterns": [
            r"(?:单位名称[：:]|[名称][：:]\s*)(.+?)(?:\s|$)"
        ],
        # 从标题行提取单位名（unit_name_row=0 时启用）
        # 算法：枚举标题里每个后缀的所有出现位置，从该位置向左扩展（只接受 ALLOWED 里的字符），
        # 得到所有候选后取「最长」一个。能正确处理「...有限公司净月团餐—分公司」这种嵌套后缀。
        # title_unit_patterns 是高级正则逃生口：若配置非空，则按列表里第一个能 search 命中的 group(1) 直接返回。
        "title_unit_suffixes": [
            "有限公司", "股份公司", "分公司", "公司", "集团",
            "医院", "卫生院", "诊所",
            "研究院", "研究所", "学院", "大学", "学校",
            "中心", "管委会", "事业部", "处", "局"
        ],
        "title_unit_allowed_chars": r"[一-龥A-Za-z0-9（）()·\-—]",
        "title_unit_patterns": [],
        "columns": {
            "transfer_total": {
                "keywords": ["转账合计"],
                "label": "转账合计（元）"
            },
            "deduction_total": {
                "keywords": ["扣款合计", "扣款"],
                "label": "扣款合计（五险一金、单位代理费）"
            },
            "net_total": {
                "keywords": ["实发合计", "实发工资", "实发"],
                "label": "实发合计（元）"
            },
            "personal_tax": {
                "keywords": ["个税", "个人所得税"],
                "label": "个人所得税"
            },
            "adjustment": {
                "keywords": ["调差", "差额调整", "调整差额", "工伤差额", "返还差额"],
                "label": "调差"
            },
            "personal_debt": {
                "keywords": ["个人欠款"],
                "label": "个人欠款"
            },
            "personal_proxy_fee": {
                "keywords": ["个人承担代理费"],
                "label": "个人承担代理费"
            },
            "union_fee": {
                "keywords": ["扣工会会费", "工会会费"],
                "label": "扣工会会费"
            },
            "accident_insurance": {
                "keywords": ["意外险（个人承担）"],
                "label": "意外险（个人承担）"
            },
            "pay_cash": {
                "keywords": ["交纳现金", "缴纳现金"],
                "label": "交纳现金"
            },
            "service_fee": {
                "keywords": ["服务费", "代理费", "管理费"],
                "label": "服务费"
            },
            "employer_insurance": {
                "keywords": ["单位缴纳", "单位社保", "单位五险一金"],
                "label": "单位缴纳"
            }
        }
    },
    "validation": {
        "enabled": False,
        "strict": True,
        "tolerance": 0.00,
        "write_back_sheet": True,
        "write_back_sheet_name": "验证结果",
        "column_sum_checks": [],
        "row_formulas": []
    },
    "table_field": {
        "columns": [
            {"key": "report_name", "label": "报表名称"},
            {"key": "unit_name", "label": "甲方单位项目名称"},
            {"key": "transfer_total", "label": "转账合计（元）"},
            {"key": "deduction_total", "label": "扣款合计（五险一金、单位代理费）"},
            {"key": "personal_tax", "label": "个人所得税合计"},
            {"key": "adjustment", "label": "其他调整合计"},
            {"key": "personal_debt", "label": "个人欠款"},
            {"key": "personal_proxy_fee", "label": "个人承担代理费"},
            {"key": "union_fee", "label": "扣工会会费"},
            {"key": "accident_insurance", "label": "意外险（个人承担）"},
            {"key": "pay_cash", "label": "交纳现金"},
            {"key": "net_total", "label": "实发合计（元）"}
        ]
    },
    "ui": {
        "template_name": "工资发放审批",
        "description": "请上传 Excel 工资表，系统将自动解析数据并提交飞书 OA 审批流程。"
    }
}


def load_config():
    config_path = os.path.join(os.path.dirname(__file__), "config.json")
    if not os.path.exists(config_path):
        return DEFAULT_CONFIG
    try:
        with open(config_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError):
        return DEFAULT_CONFIG


CONFIG = load_config()


class FeishuClient:
    def __init__(self):
        app_id = os.getenv("FEISHU_APP_ID")
        app_secret = os.getenv("FEISHU_APP_SECRET")
        if not app_id or not app_secret:
            raise ValueError("FEISHU_APP_ID and FEISHU_APP_SECRET must be set in .env")
        self.client = lark.Client.builder() \
            .app_id(app_id) \
            .app_secret(app_secret) \
            .build()
        self.file_client = lark.Client.builder() \
            .app_id(app_id) \
            .app_secret(app_secret) \
            .domain("https://open.feishu.cn") \
            .build()

    def query_user_by_mobile(self, mobile):
        """手机号查询用户 open_id"""
        request = BatchGetIdUserRequest.builder() \
            .user_id_type("open_id") \
            .request_body(BatchGetIdUserRequestBody.builder()
                          .mobiles([mobile])
                          .include_resigned(False)
                          .build()) \
            .build()
        response = self.client.contact.v3.user.batch_get_id(request)
        if not response.success():
            print(f"batch_get_id failed: code={response.code}, msg={response.msg}")
            return None
        user_list = response.data.user_list
        if not user_list:
            return None
        return user_list[0].user_id  # 当 user_id_type=open_id 时，user_id 字段的值就是 open_id

    def get_user_department(self, open_id):
        """通过 basic_batch 获取用户姓名，返回 (department_id, name)"""
        request = BasicBatchUserRequest.builder() \
            .user_id_type("open_id") \
            .request_body(BasicBatchUserRequestBody.builder()
                          .user_ids([open_id])
                          .build()) \
            .build()
        response = self.client.contact.v3.user.basic_batch(request)
        if not response.success():
            logger.error("basic_batch failed: code=%s, msg=%s, log_id=%s",
                         response.code, response.msg, response.get_log_id())
            return None, None
        users = response.data.users or []
        if not users:
            logger.warning("basic_batch returned empty for open_id=%s", open_id)
            return None, None
        user = users[0]
        name = getattr(user, "name", "") or ""
        logger.info("basic_batch: name=%s", name)
        return "", name

    def upload_file_to_feishu(self, file_path, filename=None):
        """上传文件到飞书审批系统，返回 file_code"""
        filename = filename or os.path.basename(file_path)
        with open(file_path, "rb") as f:
            file_bytes = f.read()

        multi = MultipartEncoder(fields=[
            ("name", filename),
            ("type", "attachment"),
            ("content", (filename, file_bytes, "application/octet-stream")),
        ])

        request = lark.BaseRequest.builder() \
            .http_method(lark.HttpMethod.POST) \
            .uri("/approval/openapi/v2/file/upload") \
            .headers({"Content-Type": multi.content_type}) \
            .token_types({lark.AccessTokenType.TENANT}) \
            .body(multi) \
            .build()

        response = self.file_client.request(request)
        if response.raw.status_code != 200:
            print(f"upload failed: HTTP {response.raw.status_code}")
            return None
        try:
            payload = json.loads(response.raw.content)
        except Exception as e:
            print(f"upload failed: JSON parse error: {e}")
            return None
        if payload.get("code") != 0:
            print(f"upload failed: code={payload.get('code')}, msg={payload.get('msg')}")
            return None
        return payload["data"]["code"]  # file_code (UUID)

    def build_feishu_form(self, parsed_list, title, file_codes, remark=""):
        """
        构建飞书审批表单 JSON 字符串

        parsed_list: 从 parse_excel 返回的列表，每项是一个 dict（含 report_name, unit_name, transfer_total 等）
        title: 审批标题字符串
        file_codes: 上传文件返回的 file_code 列表（UUID 字符串列表）
        remark: 备注字符串

        返回: JSON 序列化后的字符串（飞书 API 要求 form 是字符串而非数组对象）
        """
        config = load_config()
        feishu_form = config.get("feishu_form", {})
        controls_cfg = feishu_form.get("controls", {})
        table_sub_cfg = feishu_form.get("table_sub_controls", {})

        form = []

        title_cfg = controls_cfg.get("title", {})
        if title_cfg.get("feishu_id"):
            form.append({
                "id": title_cfg["feishu_id"],
                "type": "input",
                "value": title
            })

        attach_cfg = controls_cfg.get("attachments", {})
        if attach_cfg.get("feishu_id") and file_codes:
            form.append({
                "id": attach_cfg["feishu_id"],
                "type": "attachmentV2",
                "value": file_codes
            })

        table_cfg = controls_cfg.get("table", {})
        if table_cfg.get("feishu_id") and parsed_list:
            table_rows = []
            for item in parsed_list:
                row = []
                for key, sub_cfg in table_sub_cfg.items():
                    feishu_id = sub_cfg.get("feishu_id")
                    if not feishu_id:
                        continue
                    val = item.get(key, "")
                    feishu_type = sub_cfg.get("feishu_type", "input")
                    if feishu_type in ("number", "amount"):
                        val = str(float(val)) if val != "" else "0"
                    else:
                        val = str(val)
                    row.append({
                        "id": feishu_id,
                        "type": feishu_type,
                        "value": val
                    })
                table_rows.append(row)

            form.append({
                "id": table_cfg["feishu_id"],
                "type": "fieldList",
                "value": table_rows
            })

        remark_cfg = controls_cfg.get("remark", {})
        if remark_cfg.get("feishu_id"):
            form.append({
                "id": remark_cfg["feishu_id"],
                "type": remark_cfg.get("feishu_type", "input"),
                "value": remark
            })

        return json.dumps(form, ensure_ascii=False)

    def create_approval_instance(self, open_id, department_id, form_json, approval_code):
        """
        创建飞书审批实例

        返回: instance_code (字符串) 或 None（失败时）
        """
        body_builder = InstanceCreate.builder() \
            .approval_code(approval_code) \
            .open_id(open_id)
        if department_id:
            body_builder = body_builder.department_id(department_id)
        body = body_builder \
            .form(form_json) \
            .uuid(str(uuid.uuid4())) \
            .build()

        logger.info("create_approval_instance: open_id=%s, dept_id=%s",
                    open_id, department_id or "(none)")
        request = CreateInstanceRequest.builder() \
            .request_body(body) \
            .build()
        response = self.client.approval.v4.instance.create(request)

        if not response.success():
            logger.error("create instance failed: code=%s, msg=%s, log_id=%s",
                         response.code, response.msg, response.get_log_id())
            return None

        instance_code = response.data.instance_code
        logger.info("create instance success: instance_code=%s", instance_code)
        return instance_code


def _extract_year_month(text, current_year=None):
    """
    从一段文本里抽取「YYYY年MM月」并归一化为 4位年+2位月 字符串。

    支持的形态（按从严到宽顺序）：
      1) 4位年+1或2位月，连体：  2026年5月  / 2026年05月
      2) 2位年+1或2位月，连体：  26年5月
      3) 分离形态：先找最右一个「YYYY年」或「YY年」，再在其后找最近的「N月」
         例：「...2026年派遣员工5月工资明细表」

    2 位年归一化规则：>= (current_year-50) 的两位数算 21 世纪，否则算 20 世纪。
    匹配不到返回 ""。
    """
    if not text:
        return ""
    if current_year is None:
        current_year = datetime.now().year

    def normalize(y, mo):
        y = int(y); mo = int(mo)
        if y < 100:
            # 2 位年补全：在 [current-50, current+49] 范围里选
            century_base = (current_year // 100) * 100  # 2000
            cand_new = century_base + y                 # 2026
            cand_old = cand_new - 100                   # 1926
            # 选离当前年最近、且距离 < 50 的那个
            if abs(cand_new - current_year) < 50:
                y = cand_new
            else:
                y = cand_old
        if not (1 <= mo <= 12):
            return ""
        return f"{y:04d}年{mo:02d}月"

    # 规则 1：4 位年连体
    m = re.search(r"(\d{4})年(\d{1,2})月", text)
    if m:
        out = normalize(m.group(1), m.group(2))
        if out:
            return out

    # 规则 2：2 位年连体（注意要求年前面不是数字，避免吃掉 4 位年的尾巴）
    m = re.search(r"(?<!\d)(\d{2})年(\d{1,2})月", text)
    if m:
        out = normalize(m.group(1), m.group(2))
        if out:
            return out

    # 规则 3：分离形态——先找最后一个 N年，再在其后找 N月
    year_match = None
    for ym in re.finditer(r"(?<!\d)(\d{4}|\d{2})年", text):
        year_match = ym
    if year_match:
        tail = text[year_match.end():]
        mo_match = re.search(r"(\d{1,2})月", tail)
        if mo_match:
            out = normalize(year_match.group(1), mo_match.group(1))
            if out:
                return out
    return ""


def _get_excel_sheet_info(file_bytes, filename):
    """检测 Excel 文件的 sheet 数量和名称。返回 (nsheets, sheet_names)。"""
    filename_lower = filename.lower()
    if filename_lower.endswith('.xls') and not filename_lower.endswith('.xlsx'):
        import xlrd
        wb = xlrd.open_workbook(file_contents=file_bytes)
        return wb.nsheets, wb.sheet_names()
    else:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        return len(wb.sheetnames), wb.sheetnames


def _read_sheet_rows(file_bytes, filename, sheet_name=None, sheet_index=0):
    """从 Excel 文件中读取指定 sheet 的行数据。"""
    filename_lower = filename.lower()
    if filename_lower.endswith('.xls') and not filename_lower.endswith('.xlsx'):
        import xlrd
        wb = xlrd.open_workbook(file_contents=file_bytes)
        if sheet_name:
            ws = wb.sheet_by_name(sheet_name)
        else:
            ws = wb.sheet_by_index(sheet_index)
        return [ws.row_values(r) for r in range(ws.nrows)]
    else:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.worksheets[sheet_index] if sheet_index < len(wb.worksheets) else wb.active
        return list(ws.iter_rows(values_only=True))


def _extract_sheet_to_xlsx(file_bytes, filename, sheet_name, sheet_index=0):
    """从 .xls/.xlsx 文件中提取单个 sheet，保存为独立 .xlsx 字节流。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = sheet_name

    rows = _read_sheet_rows(file_bytes, filename,
                            sheet_name=sheet_name, sheet_index=sheet_index)
    for r, row in enumerate(rows, start=1):
        for c, val in enumerate(row, start=1):
            ws_out.cell(row=r, column=c, value=val)

    out = BytesIO()
    wb_out.save(out)
    return out.getvalue()


def parse_excel(file_bytes, filename, sheet_name=None, sheet_index=0):
    """
    Extract payroll data from Excel bytes.
    Returns dict with report_name, unit_name, year_month,
    transfer_total, deduction_total, net_total, tax_and_others.
    """
    rows = _read_sheet_rows(file_bytes, filename,
                            sheet_name=sheet_name, sheet_index=sheet_index)
    if not rows:
        return None

    excel_cfg = CONFIG.get("excel", DEFAULT_CONFIG["excel"])
    default_excel = DEFAULT_CONFIG["excel"]
    title_row_idx = int(excel_cfg.get("title_row", default_excel["title_row"])) - 1
    unit_name_row_cfg = int(excel_cfg.get("unit_name_row", default_excel["unit_name_row"]))
    header_start_idx = int(excel_cfg.get("header_start_row", default_excel["header_start_row"])) - 1
    header_row_count = int(excel_cfg.get("header_row_count", default_excel["header_row_count"]))

    # Title row: report name (first non-empty cell)
    report_name = ""
    if 0 <= title_row_idx < len(rows):
        for cell in rows[title_row_idx]:
            if cell is not None and str(cell).strip():
                report_name = str(cell).strip()
                break

    patterns = excel_cfg.get("unit_name_patterns", default_excel["unit_name_patterns"])
    title_patterns = excel_cfg.get("title_unit_patterns", default_excel["title_unit_patterns"])
    title_suffixes = excel_cfg.get("title_unit_suffixes", default_excel["title_unit_suffixes"])
    title_allowed = excel_cfg.get("title_unit_allowed_chars", default_excel["title_unit_allowed_chars"])

    # Unit name: from unit_name_row if configured, else extract from title row
    unit_name = ""
    # 如果 unit_name_row 提取的结果包含这些表头关键字，说明那不是单位名行，应丢弃
    _HEADER_KEYWORDS = {"序号", "姓名", "身份证号", "基本工资", "应发工资",
                        "扣款合计", "实发工资", "实发合计", "转款合计", "转账合计"}
    if unit_name_row_cfg > 0:
        unit_row_idx = unit_name_row_cfg - 1
        if 0 <= unit_row_idx < len(rows):
            row_text = " ".join(
                str(cell).strip() for cell in rows[unit_row_idx] if cell is not None
            )
            matched = False
            for pattern in patterns:
                m = re.search(pattern, row_text)
                if m:
                    unit_name = m.group(1).strip()
                    matched = True
                    break
            if not matched:
                # Fallback: 用正则严格匹配"单位名称：xxx"或"名称：xxx"模式，
                # 避免把列头中的"单位代理费""单位缴纳五险一金"误提取为单位名称
                name_unit_pat = re.compile(r'(?:单位)?名称\s*[：:]\s*(.+)')
                for cell in rows[unit_row_idx]:
                    if cell is not None:
                        val = str(cell).strip()
                        m = name_unit_pat.search(val)
                        if m:
                            unit_name = m.group(1).strip()
                            break
            # 验证提取结果：如果包含表头关键字，丢弃，从标题行提取
            if unit_name and any(kw in unit_name for kw in _HEADER_KEYWORDS):
                unit_name = ""
            # 如果提取结果为空或太短（<2字），也丢弃，从标题行提取
            if len(unit_name) < 2:
                unit_name = ""
    if not unit_name:
        # No dedicated unit row (or extraction failed) → extract from title row
        # 1) 高级用户可通过 title_unit_patterns 提供自定义正则（取第一个命中的 group 1）
        if title_patterns:
            for pattern in title_patterns:
                try:
                    m = re.search(pattern, report_name)
                except re.error:
                    continue
                if m and m.lastindex:
                    unit_name = m.group(1).strip()
                    break
        # 2) 否则用「后缀枚举 + 向左贪婪扩展 + 取最长」算法
        if not unit_name and title_suffixes:
            try:
                allowed_re = re.compile(title_allowed)
            except re.error:
                allowed_re = None
            if allowed_re is not None:
                candidates = []
                for suf in title_suffixes:
                    for m in re.finditer(re.escape(suf), report_name):
                        start = m.start()
                        while start > 0 and allowed_re.fullmatch(report_name[start - 1]):
                            start -= 1
                        cand = report_name[start : m.end()].strip()
                        if cand:
                            candidates.append(cand)
                if candidates:
                    unit_name = max(candidates, key=len)

    # Year month: 优先取标题（审计权威来源），文件名作为兜底
    # 同时单独保留两路结果，供 UI 做「标题 vs 文件名」一致性提醒
    year_month_from_title = _extract_year_month(report_name)
    year_month_from_filename = _extract_year_month(filename)
    year_month = year_month_from_title or year_month_from_filename

    summary_marker = excel_cfg.get("summary_row_marker", default_excel["summary_row_marker"])
    # Strip ALL whitespace (incl. internal) to tolerate variants like "合 计" / " 合计 "
    marker_normalized = re.sub(r"\s+", "", summary_marker)
    summary_row = None
    summary_row_idx = -1
    for ridx, row in enumerate(rows):
        if not row:
            continue
        # 先尝试第 0 列（A列）
        if row[0] is not None:
            cell = re.sub(r"\s+", "", str(row[0]))
            if cell == marker_normalized:
                summary_row = row
                summary_row_idx = ridx
                break
        # 第 0 列没找到，扩展到第 1 列（B列），仅此两列
        if summary_row is None and len(row) > 1 and row[1] is not None:
            cell = re.sub(r"\s+", "", str(row[1]))
            if cell == marker_normalized:
                summary_row = row
                summary_row_idx = ridx
                break

    if summary_row is None:
        fallback = {
            "report_name": report_name,
            "unit_name": unit_name,
            "year_month": year_month,
            "year_month_from_title": year_month_from_title,
            "year_month_from_filename": year_month_from_filename,
            "transfer_total": "0.00",
            "deduction_total": "0.00",
            "net_total": "0.00",
            "tax_and_others": "0.00",
            "column_indices": {},
            "summary_row": None,
            "data_rows": [],
            "extra_summary": {},
        }
        # 兜底也要把 excel.columns 里所有列填上 "0.00"，避免下游 KeyError
        for col_key in excel_cfg.get("columns", {}):
            fallback.setdefault(col_key, "0.00")
        return fallback

    # Header rows from config (1-based start, N rows)
    header_rows = rows[header_start_idx : header_start_idx + header_row_count]
    # 后备搜索范围：从标题行之后到合计行之前的所有行
    # 当配置的列头范围与实际文件不匹配时用此兜底，保证列关键字总能被找到
    fallback_search_start = title_row_idx + 1
    fallback_search_end = summary_row_idx
    fallback_rows = rows[fallback_search_start:fallback_search_end] if fallback_search_end > fallback_search_start else []

    def find_col_index(keywords):
        for ridx, hrow in enumerate(header_rows):
            for cidx, cell in enumerate(hrow):
                if cell is None:
                    continue
                text = str(cell).strip()
                if not text:
                    continue
                for kw in keywords:
                    if kw in text:
                        return cidx
        return -1

    def find_col_index_broad(keywords):
        """搜索所有行（从标题后到合计前），不受配置的列头范围限制"""
        for ridx, hrow in enumerate(fallback_rows):
            for cidx, cell in enumerate(hrow):
                if cell is None:
                    continue
                text = str(cell).strip()
                if not text:
                    continue
                for kw in keywords:
                    if kw in text:
                        return cidx
        return -1

    excel_cols = excel_cfg["columns"]

    # 把所有 excel.columns 列都找出索引：先用配置的列头范围搜，搜不到则用宽范围兜底
    column_indices = {}
    for col_key, col_def in excel_cols.items():
        idx = -1
        for kw in col_def.get("keywords", []):
            idx = find_col_index([kw])
            if idx != -1:
                break
        if idx == -1 and fallback_rows:
            for kw in col_def.get("keywords", []):
                idx = find_col_index_broad([kw])
                if idx != -1:
                    break
        column_indices[col_key] = idx

    # ---- 解析额外汇总行（溢缴款抵扣、甲方转款等，位于合计行之后） ----
    extra_summary_cfg = excel_cfg.get("extra_summary_rows", {})
    extra_summary_values = {}
    if extra_summary_cfg and summary_row_idx >= 0:
        xfer_idx = column_indices.get("transfer_total", -1)
        for eridx in range(summary_row_idx + 1, len(rows)):
            erow = rows[eridx]
            if not erow:
                continue
            cell0 = str(erow[0]).strip() if erow[0] is not None else ''
            cell1 = str(erow[1]).strip() if len(erow) > 1 and erow[1] is not None else ''
            for key, defn in extra_summary_cfg.items():
                for kw in defn.get("keywords", []):
                    if kw in cell0 or kw in cell1:
                        if 0 <= xfer_idx < len(erow):
                            try:
                                v = erow[xfer_idx]
                                extra_summary_values[key] = f"{float(v):.2f}" if v is not None else "0.00"
                            except (ValueError, TypeError):
                                extra_summary_values[key] = "0.00"
                        else:
                            extra_summary_values[key] = "0.00"
                        break
    # 数据行起始位置智能检测：从标题行向下扫描，找到第1列为数字序号的行作为数据起点
    # 这样即使 header_start_row 配置与实际文件不匹配，也能正确切分
    data_start_idx = None
    for ridx in range(title_row_idx + 1, summary_row_idx):
        r = rows[ridx]
        if r and len(r) > 0:
            v = r[0]
            if isinstance(v, (int, float)):
                data_start_idx = ridx
                break
            if isinstance(v, str) and re.match(r'^\d+$', v.strip()):
                data_start_idx = ridx
                break
    if data_start_idx is not None:
        header_end_idx = data_start_idx
    else:
        header_end_idx = header_start_idx + header_row_count

    # 数据行 = 表头之后 到 合计行之前；过滤掉全空行
    data_rows = []
    for r in rows[header_end_idx:summary_row_idx]:
        if r is None:
            continue
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in r):
            continue
        data_rows.append(r)

    def _sum_data_rows(col_key):
        """从数据行汇总某列的总和，供合计行缺失时兜底"""
        cidx = column_indices.get(col_key, -1)
        if cidx < 0 or not data_rows:
            return 0.0
        total = 0.0
        for row in data_rows:
            if cidx < len(row):
                v = row[cidx]
                try:
                    total += float(v) if v is not None else 0.0
                except (ValueError, TypeError):
                    pass
        return round(total, 2)

    def get_val(idx, col_key=None):
        if idx >= 0 and idx < len(summary_row):
            v = summary_row[idx]
            if v is not None:
                try:
                    return f"{float(v):.2f}"
                except (ValueError, TypeError):
                    return str(v).strip() or "0.00"
        # 合计行单元格为空 → 从数据行汇总兜底
        if col_key and data_rows:
            s = _sum_data_rows(col_key)
            return f"{s:.2f}"
        return "0.00"

    transfer_idx = column_indices.get("transfer_total", -1)
    deduction_idx = column_indices.get("deduction_total", -1)
    net_idx = column_indices.get("net_total", -1)

    transfer_total = get_val(transfer_idx, "transfer_total")
    deduction_total = get_val(deduction_idx, "deduction_total")
    net_total = get_val(net_idx, "net_total")

    # 把 excel.columns 里所有定义过的列，都从合计行取值（合计行缺失则汇总数据行），统一加到返回 dict
    column_summary_values = {}
    for col_key, cidx in column_indices.items():
        column_summary_values[col_key] = get_val(cidx, col_key)

    # 额外汇总行缺失时兜底：party_a_transfer 默认等于转款合计，其余默认为 0
    xfer_default = column_summary_values.get("transfer_total", "0.00")
    for ekey in extra_summary_cfg:
        if ekey not in extra_summary_values:
            extra_summary_values[ekey] = xfer_default if ekey == "party_a_transfer" else "0.00"

    try:
        tax_val = float(transfer_total) - float(deduction_total) - float(net_total)
        ptax_key = "personal_tax"
        if ptax_key in column_summary_values:
            tax_val -= float(column_summary_values[ptax_key])
        if abs(tax_val) < 0.005:
            tax_val = 0.0
        tax_and_others = f"{tax_val:.2f}"
    except (ValueError, TypeError):
        tax_and_others = "0.00"

    result = {
        "report_name": report_name,
        "unit_name": unit_name,
        "year_month": year_month,
        "year_month_from_title": year_month_from_title,
        "year_month_from_filename": year_month_from_filename,
        "tax_and_others": tax_and_others,
        "column_indices": column_indices,
        "summary_row": summary_row,
        "data_rows": data_rows,
        "extra_summary": extra_summary_values,
        "_cfg_cols": {
            "header_start_row": excel_cfg.get("header_start_row", default_excel["header_start_row"]),
            "header_row_count": excel_cfg.get("header_row_count", default_excel["header_row_count"]),
        },
    }
    # 把 excel.columns 里所有列的合计值都加上（包括 transfer/deduction/net）
    # 这样 table_field 可以引用任意已配置的列
    result.update(column_summary_values)
    result.update(extra_summary_values)
    return result


def _to_money(v):
    """把单元格值转成 round 到 2 位小数的 float；None/空/非数 → 0.0。"""
    if v is None:
        return 0.0
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return 0.0
        try:
            return round(float(s), 2)
        except ValueError:
            return 0.0
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return 0.0


def _row_label(row, default_label):
    """数据行标识：序号+姓名 → 'row 8 张吉'；否则只用 default_label。"""
    if not row:
        return default_label
    first = row[0]
    name = row[1] if len(row) > 1 else None
    if isinstance(first, (int, float)) and name and str(name).strip():
        return f"{default_label} {str(name).strip()}"
    return default_label


def validate_payroll(parsed, validation_cfg, excel_cols):
    """
    校验解析后的工资表数据，返回结构化结果。
    parsed: parse_excel 的返回值（必须含 column_indices/summary_row/data_rows）
    validation_cfg: CONFIG['validation']
    excel_cols: CONFIG['excel']['columns']

    校验类型：
      column_sum:           sum(数据列) == 合计行该列
      row_formula_summary:  合计行 lhs == sum(rhs_plus) - sum(rhs_minus)
      row_formula_rows:     每个数据行同公式
    """
    tolerance = float(validation_cfg.get("tolerance", 0.0))
    column_indices = parsed.get("column_indices", {})
    summary_row = parsed.get("summary_row")
    data_rows = parsed.get("data_rows", [])

    checks = []

    def col_label(key):
        if key in excel_cols and excel_cols[key].get("label"):
            return excel_cols[key]["label"]
        return key

    def col_value(row, key):
        """从 row 中按 column_indices 取某 key 列的值（→ float）。"""
        idx = column_indices.get(key, -1)
        if idx < 0 or row is None or idx >= len(row):
            return 0.0
        return _to_money(row[idx])

    def require_col(key, where):
        if key not in excel_cols:
            raise ValueError(
                f"validation 配置中 {where} 引用了 '{key}'，但 excel.columns 未定义该列"
            )
        if column_indices.get(key, -1) < 0:
            # 列已声明但未在 Excel 表头中找到——这是数据/配置不匹配，不阻断校验
            # 取值会按 0 计，会触发对应的失败 issue（用户能从结果看出来）
            pass

    # === A. 纵向列加总 ===
    for spec in validation_cfg.get("column_sum_checks", []) or []:
        col_key = spec.get("column")
        if not col_key:
            continue
        require_col(col_key, "column_sum_checks")
        col_sum = sum(col_value(r, col_key) for r in data_rows)
        col_sum = round(col_sum, 2)
        summary_val = col_value(summary_row, col_key)
        diff = round(col_sum - summary_val, 2)
        passed = abs(diff) <= tolerance
        checks.append({
            "kind": "column_sum",
            "name": f"{col_label(col_key)} 列加总",
            "column_label": col_label(col_key),
            "col_sum": col_sum,
            "summary": summary_val,
            "diff": diff,
            "passed": passed,
            "detail": "" if passed else f"列加总 {col_sum:.2f} 与合计行 {summary_val:.2f} 相差 {diff:+.2f}"
        })

    # === B+C. 横向公式 ===
    for formula in validation_cfg.get("row_formulas", []) or []:
        name = formula.get("name", "<未命名公式>")
        lhs = formula.get("lhs")
        rhs_plus = formula.get("rhs_plus", []) or []
        rhs_minus = formula.get("rhs_minus", []) or []
        if not lhs:
            continue
        require_col(lhs, f"row_formulas[{name}].lhs")
        for k in rhs_plus:
            require_col(k, f"row_formulas[{name}].rhs_plus")
        for k in rhs_minus:
            require_col(k, f"row_formulas[{name}].rhs_minus")

        def rhs_of(row):
            return round(
                sum(col_value(row, k) for k in rhs_plus)
                - sum(col_value(row, k) for k in rhs_minus),
                2
            )

        # B: 合计行
        L = col_value(summary_row, lhs)
        R = rhs_of(summary_row)
        diff = round(L - R, 2)
        passed = abs(diff) <= tolerance
        checks.append({
            "kind": "row_formula_summary",
            "name": f"{name} (合计行)",
            "lhs_value": L,
            "rhs_value": R,
            "diff": diff,
            "passed": passed,
            "detail": "" if passed else f"合计行 lhs={L:.2f}，rhs={R:.2f}，差 {diff:+.2f}"
        })

        # C: 每个数据行
        failed_rows = []
        for ridx, r in enumerate(data_rows):
            L_r = col_value(r, lhs)
            R_r = rhs_of(r)
            d = round(L_r - R_r, 2)
            if abs(d) > tolerance:
                failed_rows.append({
                    "row_label": _row_label(r, f"第 {ridx + 1} 行"),
                    "lhs": L_r,
                    "rhs": R_r,
                    "diff": d,
                })
        total_rows = len(data_rows)
        passed_rows = total_rows - len(failed_rows)
        passed = len(failed_rows) == 0
        if passed:
            detail = f"{passed_rows}/{total_rows} 行通过"
        else:
            sample = "; ".join(
                f"{f['row_label']} 差 {f['diff']:+.2f}" for f in failed_rows[:3]
            )
            more = f"...（共 {len(failed_rows)} 行不通过）" if len(failed_rows) > 3 else ""
            detail = f"{passed_rows}/{total_rows} 行通过；失败例: {sample}{more}"
        checks.append({
            "kind": "row_formula_rows",
            "name": f"{name} (每行)",
            "total_rows": total_rows,
            "passed_rows": passed_rows,
            "failed_rows": failed_rows,
            "passed": passed,
            "detail": detail
        })

    # === D. 表格格式验证：检查关键列是否在表头中找到、合计行是否有实际数值 ===
    # 当配置的 header_start_row / header_row_count 与实际文件不匹配时，
    # 即使列关键字被宽范围搜索找到，数据行和合计行也可能没有对应的值，
    # 导致全部识别为 0 但校验仍通过（因为 0=0）。这个检查专门捕获这种情况。
    _cfg_cols = parsed.get("_cfg_cols", {})
    hdr_start = int(_cfg_cols.get("header_start_row", 3))
    hdr_count = int(_cfg_cols.get("header_row_count", 3))
    hdr_end = hdr_start + hdr_count - 1
    critical_cols = ["transfer_total", "deduction_total", "net_total"]
    col_labels_map = {k: col_label(k) for k in critical_cols}

    missing_cols = [col_labels_map[k] for k in critical_cols if column_indices.get(k, -1) < 0]
    if missing_cols:
        passed = False
        detail = (
            f"未在表头中找到以下关键列：{'、'.join(missing_cols)}。"
            f"请确认工资表表头起始行为第 {hdr_start} 行，第 {hdr_start}–{hdr_end} 行为组合表头（当前配置）。"
        )
    else:
        # 列都找到了，检查合计行和数据行是否有实际数值
        has_real_data = False
        for k in critical_cols:
            idx = column_indices.get(k, -1)
            if idx >= 0 and summary_row and idx < len(summary_row):
                v = summary_row[idx]
                if v is not None:
                    try:
                        if float(v) != 0.0:
                            has_real_data = True
                            break
                    except (ValueError, TypeError):
                        pass
        if not has_real_data and data_rows:
            passed = False
            # 用 config 的实际值提示
            detail = (
                f"识别到的转账合计、扣款合计、实发合计均为 0，但存在 {len(data_rows)} 行明细数据。"
                f"可能是表格格式有误，请确认工资表表头起始行为第 {hdr_start} 行、"
                f"第 {hdr_start}–{hdr_end} 行为组合表头。如文件格式不同，请联系管理员调整配置。"
            )
        else:
            passed = True
            detail = ""

    checks.append({
        "kind": "table_format",
        "name": "表格格式验证",
        "passed": passed,
        "detail": detail,
    })

    # === F. 零金额检查：转款合计和实发合计同时为 0 则阻断 ===
    def _is_zero(val):
        if val is None:
            return True
        if isinstance(val, str):
            return val.strip() in ("", "0", "0.00")
        try:
            return float(val) == 0.0
        except (ValueError, TypeError):
            return False
    trans_val = parsed.get("transfer_total", "0.00")
    net_val = parsed.get("net_total", "0.00")
    zero_amount_ok = not (_is_zero(trans_val) and _is_zero(net_val))
    checks.append({
        "kind": "zero_amount",
        "name": "零金额检查",
        "passed": zero_amount_ok,
        "detail": "" if zero_amount_ok else "转款合计和实发合计均为 0，请确认表头配置是否正确",
    })

    # === E. 额外汇总行校验（溢缴款抵扣、甲方转款等） ===
    for spec in validation_cfg.get("extra_summary_checks", []) or []:
        name = spec.get("name", "<未命名额外汇总校验>")
        col_key = spec.get("column", "transfer_total")
        rhs_plus_keys = spec.get("rhs_plus_row_keys", [])
        rhs_minus_keys = spec.get("rhs_minus_row_keys", [])
        if not col_key:
            continue

        lhs_val = float(parsed.get(col_key, "0.00"))
        extra = parsed.get("extra_summary", {})
        rhs_val = 0.0
        for k in rhs_plus_keys:
            rhs_val += float(extra.get(k, "0.00"))
        for k in rhs_minus_keys:
            rhs_val -= float(extra.get(k, "0.00"))
        rhs_val = round(rhs_val, 2)
        diff = round(lhs_val - rhs_val, 2)
        ok = abs(diff) <= tolerance

        detail_parts = [f"{col_label(col_key)} = {lhs_val:.2f}"]
        rhs_terms = []
        for k in rhs_plus_keys:
            rhs_terms.append(f"{k}({float(extra.get(k, '0.00')):.2f})")
        for k in rhs_minus_keys:
            rhs_terms.append(f"-{k}({float(extra.get(k, '0.00')):.2f})")
        detail_parts.append(" + ".join(rhs_terms) + f" = {rhs_val:.2f}")
        detail_parts.append("差 " + f"{diff:+.2f}")
        detail = "" if ok else " | ".join(detail_parts)

        checks.append({
            "kind": "extra_summary",
            "name": name,
            "passed": ok,
            "detail": detail,
        })

    passed_count = sum(1 for c in checks if c["passed"])
    failed_count = len(checks) - passed_count
    return {
        "ok": failed_count == 0,
        "passed_count": passed_count,
        "failed_count": failed_count,
        "checks": checks,
    }


def _parsed_key(item):
    """从解析结果中生成唯一 key：单 sheet 用文件名，多 sheet 用 文件名::sheet名"""
    fname = item.get("filename", "")
    sname = item.get("sheet_name", "")
    if sname:
        return f"{fname}::{sname}"
    return fname


def check_signatures(file_bytes, filename, required_sigs):
    """
    Scan all cells in an Excel file for required signature keywords.

    required_sigs 支持两种格式：
      - 扁平列表 ["A", "B", "C"] → 全部必须存在（向后兼容）
      - 分组列表 [["A","B","C","D"], ["A","B","C"]] → 任一组全部匹配即通过

    Returns {"ok": bool, "found": [...], "missing": [...]}
    """
    if required_sigs and isinstance(required_sigs[0], list):
        groups = required_sigs
    else:
        groups = [required_sigs] if required_sigs else []

    filename_lower = filename.lower()
    if filename_lower.endswith('.xls') and not filename_lower.endswith('.xlsx'):
        import xlrd
        wb = xlrd.open_workbook(file_contents=file_bytes)
        ws = wb.sheet_by_index(0)
        all_text = " ".join(
            str(ws.cell_value(r, c)) for r in range(ws.nrows) for c in range(ws.ncols)
        )
    else:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        ws = wb.active
        all_text = " ".join(
            str(cell.value or "") for row in ws.iter_rows() for cell in row
        )

    all_text_lower = all_text.lower()

    def kw_matches(kw):
        return any(
            alias.strip().lower() in all_text_lower
            for alias in kw.split("|")
        )

    for group in groups:
        missing_in_group = [kw for kw in group if not kw_matches(kw)]
        if not missing_in_group:
            return {"ok": True, "found": group, "missing": []}

    best = min(groups, key=lambda g: sum(1 for kw in g if not kw_matches(kw)))
    found = [kw for kw in best if kw_matches(kw)]
    missing = [kw for kw in best if not kw_matches(kw)]
    return {"ok": False, "found": found, "missing": missing}


def append_validation_sheet(file_bytes, validation_result,
                            sheet_name="验证结果",
                            source_filename=""):
    """
    在 Excel 字节流末尾追加一个 sheet，写入校验结果。原表数据不动。
    如果同名 sheet 已存在（重复处理场景），先删后建，保证幂等。
    返回新的字节流。

    注：.xls (Excel 97-2003) 格式为只读，不支持追加 sheet，
    此时返回原 bytes，由调用方提示用户。
    """
    from openpyxl.styles import PatternFill, Font, Alignment

    wb = openpyxl.load_workbook(BytesIO(file_bytes))
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True)
    meta_font = Font(italic=True, color="666666")
    pass_fill = PatternFill("solid", fgColor="E6F4EA")  # 淡绿
    fail_fill = PatternFill("solid", fgColor="FCE8E6")  # 淡红
    header_fill = PatternFill("solid", fgColor="EFEFEF")
    center = Alignment(horizontal="center", vertical="center")

    # 标题
    ws["A1"] = "工资表校验结果"
    ws["A1"].font = title_font
    ws.merge_cells("A1:D1")
    ws["A1"].alignment = center

    # 元数据
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    passed = validation_result.get("passed_count", 0)
    failed = validation_result.get("failed_count", 0)
    meta_rows = [
        ("生成时间", now_str),
        ("源文件", source_filename or "(未提供)"),
        ("汇总", f"{passed} 项通过 / {failed} 项失败"),
    ]
    for i, (k, v) in enumerate(meta_rows, start=2):
        ws.cell(row=i, column=1, value=k).font = meta_font
        ws.cell(row=i, column=2, value=v).font = meta_font

    # 明细表头
    head_row = 2 + len(meta_rows) + 1  # 留一行空
    headers = ["校验项", "类型", "结果", "说明"]
    for cidx, h in enumerate(headers, start=1):
        c = ws.cell(row=head_row, column=cidx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    kind_label = {
        "column_sum": "纵向加总",
        "row_formula_summary": "横向公式",
        "row_formula_rows": "横向公式",
        "table_format": "表格格式",
        "extra_summary": "额外汇总",
        "zero_amount": "零金额检查",
    }

    # 明细行
    for i, ck in enumerate(validation_result.get("checks", []), start=head_row + 1):
        ws.cell(row=i, column=1, value=ck.get("name", ""))
        ws.cell(row=i, column=2, value=kind_label.get(ck.get("kind"), ck.get("kind", "")))
        ws.cell(row=i, column=3, value="✅ 通过" if ck.get("passed") else "❌ 失败")
        ws.cell(row=i, column=4, value=ck.get("detail", ""))
        fill = pass_fill if ck.get("passed") else fail_fill
        for cidx in range(1, 5):
            ws.cell(row=i, column=cidx).fill = fill

    # 列宽
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 60

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def build_summary_workbook(parsed_list, validation_results, tf_columns,
                           write_back_sheet_name="验证结果",
                           signature_results=None,
                           required_signatures=None):
    """
    生成「工资发放汇总表」xlsx，包含两个 sheet：
      1) 汇总数据：每行一个附件，列同数据预览（文件名、年月、报表字段...、验证结果）
      2) 验证明细：所有附件的校验结果合并，多一列"附件名"区分

    解决两个问题：
      - .xls 附件无法在原文件内回写验证结果
      - 多附件场景下没有统一的"全局视图"
    """
    from openpyxl.styles import PatternFill, Font, Alignment

    wb = openpyxl.Workbook()

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True)
    meta_font = Font(italic=True, color="666666")
    header_fill = PatternFill("solid", fgColor="EFEFEF")
    pass_fill = PatternFill("solid", fgColor="E6F4EA")
    fail_fill = PatternFill("solid", fgColor="FCE8E6")
    center = Alignment(horizontal="center", vertical="center")
    money_font = Font(name="Consolas")

    # ===== Sheet 1: 汇总数据 =====
    ws = wb.active
    ws.title = "汇总数据"

    headers = ["文件名", "年月"] + [c["label"] for c in tf_columns] + ["验证结果"]
    ws["A1"] = "工资发放汇总数据"
    ws["A1"].font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"].alignment = center

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A2"] = f"生成时间：{now_str}    附件数：{len(parsed_list)}"
    ws["A2"].font = meta_font
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))

    head_row = 4
    for cidx, h in enumerate(headers, start=1):
        c = ws.cell(row=head_row, column=cidx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    for i, p in enumerate(parsed_list, start=head_row + 1):
        ws.cell(row=i, column=1, value=p.get("filename", ""))
        ws.cell(row=i, column=2, value=p.get("year_month", ""))
        for cidx, col in enumerate(tf_columns, start=3):
            v = p.get(col["key"], "")
            # 金额列尝试转 float，方便 Excel 求和/排序
            if isinstance(v, str):
                try:
                    v_num = float(v)
                    cell = ws.cell(row=i, column=cidx, value=v_num)
                    cell.number_format = "#,##0.00"
                    cell.font = money_font
                except ValueError:
                    ws.cell(row=i, column=cidx, value=v)
            else:
                ws.cell(row=i, column=cidx, value=v)
        # 验证结果列
        vr = validation_results.get(_parsed_key(p))
        sr = (signature_results or {}).get(p.get("filename"))
        status_parts = []
        fill = None

        if sr is not None:
            if sr["ok"]:
                status_parts.append("✅ 签名栏齐全")
            else:
                missing = "、".join(sr["missing"])
                status_parts.append(f"❌ 缺少签名栏: {missing}")
                fill = fail_fill

        if vr is None:
            if not status_parts:
                status = "未启用"
                fill = None
            else:
                status = " | ".join(status_parts)
        elif vr["ok"]:
            status_parts.append(f"✅ 数值校验全部通过 ({vr['passed_count']} 项)")
            if fill is None:
                fill = pass_fill
            status = " | ".join(status_parts)
        else:
            status_parts.append(f"⚠️ {vr['failed_count']} 项未通过 (共 {vr['passed_count']+vr['failed_count']} 项)")
            fill = fail_fill
            status = " | ".join(status_parts)
        if not status_parts and vr is None and sr is None:
            status = "未启用"
        status_cell = ws.cell(row=i, column=len(headers), value=status)
        if fill is not None:
            status_cell.fill = fill

    # 末尾追加一行"总计"，对所有金额列求和（仅数值列）
    if parsed_list:
        total_row = head_row + 1 + len(parsed_list)
        ws.cell(row=total_row, column=1, value="总计").font = header_font
        for cidx, col in enumerate(tf_columns, start=3):
            # 试求和：把每行该列取出，能转 float 的相加
            total = 0.0
            has_num = False
            for p in parsed_list:
                v = p.get(col["key"], "")
                try:
                    total += float(v)
                    has_num = True
                except (ValueError, TypeError):
                    pass
            if has_num:
                cell = ws.cell(row=total_row, column=cidx, value=round(total, 2))
                cell.number_format = "#,##0.00"
                cell.font = Font(name="Consolas", bold=True)
                cell.fill = header_fill
            else:
                ws.cell(row=total_row, column=cidx, value="").fill = header_fill
        ws.cell(row=total_row, column=2, value="").fill = header_fill  # 年月列留空
        ws.cell(row=total_row, column=1).fill = header_fill
        ws.cell(row=total_row, column=len(headers), value="").fill = header_fill

    # 列宽
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14
    for ci in range(3, 3 + len(tf_columns)):
        ws.column_dimensions[ws.cell(row=head_row, column=ci).column_letter].width = 22
    ws.column_dimensions[ws.cell(row=head_row, column=len(headers)).column_letter].width = 36

    # ===== Sheet 2: 验证明细 =====
    ws2 = wb.create_sheet("验证明细")
    ws2["A1"] = "各附件验证明细"
    ws2["A1"].font = title_font
    ws2.merge_cells("A1:E1")
    ws2["A1"].alignment = center

    detail_headers = ["附件名", "校验项", "类型", "结果", "说明"]
    for cidx, h in enumerate(detail_headers, start=1):
        c = ws2.cell(row=3, column=cidx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    kind_label = {
        "column_sum": "纵向加总",
        "row_formula_summary": "横向公式",
        "row_formula_rows": "横向公式",
        "table_format": "表格格式",
        "extra_summary": "额外汇总",
        "zero_amount": "零金额检查",
        "signature_check": "签名栏",
    }
    cur = 4
    for p in parsed_list:
        fn = p.get("filename", "")
        vr = validation_results.get(_parsed_key(p))
        sr = (signature_results or {}).get(fn)

        # 签名栏检查行（放在最前面，最重要）
        if sr is not None:
            ws2.cell(row=cur, column=1, value=fn)
            ws2.cell(row=cur, column=2, value="签名栏检查")
            ws2.cell(row=cur, column=3, value=kind_label["signature_check"])
            if sr["ok"]:
                ws2.cell(row=cur, column=4, value="✅ 通过")
                ws2.cell(row=cur, column=5, value=f"已找到: {'、'.join(sr['found'])}")
                fill = pass_fill
            else:
                ws2.cell(row=cur, column=4, value="❌ 失败")
                detail = f"缺少: {'、'.join(sr['missing'])}"
                if sr["found"]:
                    detail += f"；已找到: {'、'.join(sr['found'])}"
                ws2.cell(row=cur, column=5, value=detail)
                fill = fail_fill
            for cidx in range(1, 6):
                ws2.cell(row=cur, column=cidx).fill = fill
            cur += 1

        if vr is None:
            ws2.cell(row=cur, column=1, value=fn)
            ws2.cell(row=cur, column=2, value="(未启用校验)").fill = header_fill
            cur += 1
            continue
        for ck in vr.get("checks", []):
            ws2.cell(row=cur, column=1, value=fn)
            ws2.cell(row=cur, column=2, value=ck.get("name", ""))
            ws2.cell(row=cur, column=3, value=kind_label.get(ck.get("kind"), ck.get("kind", "")))
            ws2.cell(row=cur, column=4, value="✅ 通过" if ck.get("passed") else "❌ 失败")
            ws2.cell(row=cur, column=5, value=ck.get("detail", ""))
            fill = pass_fill if ck.get("passed") else fail_fill
            for cidx in range(1, 6):
                ws2.cell(row=cur, column=cidx).fill = fill
            cur += 1

    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 38
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 10
    ws2.column_dimensions["E"].width = 60

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def generate_title(unit_names, year_month, amounts):
    """
    unit_names: list of unit names (ordered by amount desc)
    amounts: list of transfer amounts (same order)
    year_month: "2026年03月" format
    """
    n = len(unit_names)
    if n == 0:
        return f"{year_month}工资发放请示"
    if n == 1:
        return f"{unit_names[0]}{year_month}工资发放请示"
    if n == 2:
        return f"{unit_names[0]}、{unit_names[1]}{year_month}工资发放请示"
    # 3+ units
    return f"{unit_names[0]}、{unit_names[1]}等{n}家单位{year_month}工资发放请示"


def main():
    ui_config = CONFIG.get("ui", DEFAULT_CONFIG["ui"])
    template_name = ui_config.get("template_name", "工资发放审批")
    description = ui_config.get("description", "")

    st.title(f"📋 {template_name}")
    if description:
        st.info(description)

    if "logged_in" not in st.session_state:
        st.session_state.logged_in = False
    if "open_id" not in st.session_state:
        st.session_state.open_id = ""
    if "department_id" not in st.session_state:
        st.session_state.department_id = ""
    if "user_name" not in st.session_state:
        st.session_state.user_name = ""

    client = FeishuClient()
    query_feishu_approval_definition()

    # Section 1: Phone Login
    if not st.session_state.logged_in:
        st.subheader("手机号登录")
        mobile = st.text_input("手机号", value="")
        if st.button("登录"):
            if not mobile:
                st.error("请输入手机号")
                return
            try:
                open_id = client.query_user_by_mobile(mobile)
                if not open_id:
                    st.error("手机号未找到对应用户")
                    return
                department_id, user_name = client.get_user_department(open_id)
                if not user_name:
                    logger.warning("get_user_department returned empty name for open_id=%s, dept=%s", open_id, department_id)
                else:
                    logger.info("Logged in: %s (open_id=%s)", user_name, open_id)
                st.session_state.logged_in = True
                st.session_state.open_id = open_id
                st.session_state.department_id = department_id or ""
                st.session_state.user_name = user_name or open_id
                st.success(f"登录成功：{user_name or open_id}")
                st.rerun()
            except Exception as e:
                st.error(f"登录失败：{e}")
        return

    # Section 2: File Upload
    st.subheader(f"欢迎，{st.session_state.user_name}")
    col1, col2 = st.columns([1, 1])
    with col2:
        if st.button("退出登录"):
            for key in ["logged_in", "open_id", "department_id", "user_name"]:
                st.session_state.pop(key, None)
            st.rerun()
    uploaded_files = st.file_uploader(
        "上传工资表", type=["xlsx", "xls"], accept_multiple_files=True
    )

    if not uploaded_files:
        st.info("请上传一个或多个 Excel 工资表文件")
        return

    # Parse each file (support multi-sheet auto-split)
    parsed_list = []
    multi_sheet_files = set()
    for upfile in uploaded_files:
        file_bytes = upfile.read()
        upfile.seek(0)
        nsheets, sheet_names = _get_excel_sheet_info(file_bytes, upfile.name)
        if nsheets > 1:
            multi_sheet_files.add(upfile.name)
            for sidx, sname in enumerate(sheet_names):
                parsed = parse_excel(file_bytes, upfile.name,
                                     sheet_name=sname, sheet_index=sidx)
                if parsed:
                    extracted = _extract_sheet_to_xlsx(
                        file_bytes, upfile.name, sname, sidx,
                    )
                    parsed_list.append({
                        "filename": upfile.name,
                        "sheet_name": sname,
                        "sheet_index": sidx,
                        "_extracted_bytes": extracted,
                        **parsed,
                    })
        else:
            parsed = parse_excel(file_bytes, upfile.name)
            if parsed:
                extracted = _extract_sheet_to_xlsx(
                    file_bytes, upfile.name,
                    sheet_name="", sheet_index=0,
                )
                parsed_list.append({
                    "filename": upfile.name,
                    "sheet_name": "",
                    "sheet_index": 0,
                    "_extracted_bytes": extracted,
                    **parsed,
                })

    if multi_sheet_files:
        st.warning(
            f"⚠️ 以下文件包含多个 sheet，"
            f"已自动拆分为独立文件处理，建议将每个 sheet 保存为单独文件上传："
            f"{'、'.join(multi_sheet_files)}"
        )

    if not parsed_list:
        st.error("未能解析任何文件，请检查格式")
        return

    # 标题 vs 文件名年月一致性检查
    for p in parsed_list:
        t = p.get("year_month_from_title", "")
        f = p.get("year_month_from_filename", "")
        label = p["filename"]
        if p.get("sheet_name"):
            label += f" [{p['sheet_name']}]"
        if t and f and t != f:
            st.warning(
                f"⚠️ {label}：标题中年月「{t}」与文件名年月「{f}」不一致，"
                f"已以**标题**为准。请确认报表标题是否需要更正。"
            )
        elif not t and f:
            st.warning(
                f"⚠️ {label}：报表标题中未识别到年月，已退回使用文件名年月「{f}」。"
                f"建议在标题中明确写出年月。"
            )

    # 工资表内容校验
    val_cfg = CONFIG.get("validation", DEFAULT_CONFIG["validation"])
    required_sigs = val_cfg.get("required_signatures", [])
    validation_results = {}  # {filename: validation_result}
    signature_results = {}   # {filename: signature_check_result}
    submit_blocked = False
    signature_check_blocked = False

    if required_sigs:
        for upfile in uploaded_files:
            file_bytes = upfile.read()
            upfile.seek(0)
            result = check_signatures(file_bytes, upfile.name, required_sigs)
            signature_results[upfile.name] = result
            if not result["ok"]:
                signature_check_blocked = True
                missing_list = "、".join(result["missing"])
                st.error(
                    f"⚠️ {upfile.name}：缺少签名栏信息 — {missing_list}。"
                    f"请确保 Excel 文件中包含这些字段后再上传。"
                )
    if val_cfg.get("enabled"):
        excel_cols_def = CONFIG.get("excel", {}).get("columns", {})
        for p in parsed_list:
            key = _parsed_key(p)
            try:
                vr = validate_payroll(p, val_cfg, excel_cols_def)
            except ValueError as e:
                label = p["filename"]
                if p.get("sheet_name"):
                    label += f" [{p['sheet_name']}]"
                st.error(f"⚠️ {label} 校验配置错误：{e}")
                vr = None
                if val_cfg.get("strict"):
                    submit_blocked = True
            validation_results[key] = vr
            if vr is not None and not vr["ok"] and val_cfg.get("strict"):
                submit_blocked = True

    # Preview table
    st.subheader("数据预览")
    preview_data = []
    tf_columns = CONFIG["table_field"]["columns"]
    for p in parsed_list:
        key = _parsed_key(p)
        display_name = p["filename"]
        if p.get("sheet_name"):
            display_name += f" [{p['sheet_name']}]"
        row = {"文件名": display_name, "年月": p["year_month"]}
        for col in tf_columns:
            row[col["label"]] = p.get(col["key"], "")
        sig_parts = []

        sr = signature_results.get(p["filename"])
        if sr is not None:
            if sr["ok"]:
                sig_parts.append("✅ 签名栏齐全")
            else:
                sig_parts.append(f"❌ 缺少签名栏: {'、'.join(sr['missing'])}")

        if val_cfg.get("enabled"):
            vr = validation_results.get(key)
            if vr is None:
                sig_parts.append("⚠️ 配置错误")
            elif vr["ok"]:
                sig_parts.append(f"✅ 数值全部通过 ({vr['passed_count']} 项)")
            else:
                sig_parts.append(f"⚠️ {vr['failed_count']}/{vr['passed_count'] + vr['failed_count']} 项未通过")

        if sig_parts:
            row["验证结果"] = " | ".join(sig_parts)
        preview_data.append(row)

    # 隐藏所有文件均为 0 的列，减少横向滚动，方便用户聚焦有数值的列
    # 不影响飞书 API 提交的数据（parsed_list / tf_columns 不变）
    zero_column_labels = set()
    for col in tf_columns:
        if all(p.get(col["key"], "0.00") in ("0.00", "0", "", None) for p in parsed_list):
            zero_column_labels.add(col["label"])
    if zero_column_labels:
        for row in preview_data:
            for c in zero_column_labels:
                row.pop(c, None)
        st.caption(f"已隐藏全为零的列：{'、'.join(sorted(zero_column_labels))}。数据仍会提交至飞书。")
    st.dataframe(preview_data)

    # 生成审批标题（提前到这里，汇总表文件名要从标题派生）
    # Sort by transfer_total desc
    sorted_items = sorted(
        parsed_list,
        key=lambda x: float(x["transfer_total"]) if x["transfer_total"] else 0,
        reverse=True,
    )
    unit_names = [p["unit_name"] for p in sorted_items]
    amounts = [p["transfer_total"] for p in sorted_items]
    year_month = sorted_items[0]["year_month"] if sorted_items else ""
    title = generate_title(unit_names, year_month, amounts)

    # 生成「工资发放汇总表」xlsx：把数据预览 + 所有附件的验证明细打包成一个文件
    # 用途：(1) .xls 附件无法回写验证 sheet，靠这里集中展示；
    #      (2) 多附件场景下提供全局视图；(3) 作为额外附件随审批一起归档
    # 文件名从审批标题派生：把"工资发放请示"换成"工资发放汇总表及验证明细"
    if "工资发放请示" in title:
        summary_filename = title.replace("工资发放请示", "工资发放汇总表及验证明细") + ".xlsx"
    else:
        summary_filename = (
            f"{title}汇总表及验证明细.xlsx" if title else "工资发放汇总表及验证明细.xlsx"
        )
    try:
        summary_bytes = build_summary_workbook(
            parsed_list,
            validation_results,
            tf_columns,
            write_back_sheet_name=val_cfg.get("write_back_sheet_name", "验证结果"),
            signature_results=signature_results,
            required_signatures=required_sigs,
        )
        st.download_button(
            label=f"📥 下载汇总表（{summary_filename}）",
            data=summary_bytes,
            file_name=summary_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        summary_bytes = None
        st.warning(f"⚠️ 生成汇总表失败：{e}")

    # 提示：若有失败，告诉用户去看 Excel sheet
    if val_cfg.get("enabled") and any(
        vr is not None and not vr["ok"] for vr in validation_results.values()
    ):
        st.info(
            f"📋 详细校验明细已写入附件的「{val_cfg.get('write_back_sheet_name', '验证结果')}」sheet（仅 .xlsx 附件），"
            f"或参见上方汇总表的「验证明细」sheet。"
        )

    # 每个文件提供下载按钮
    for p in parsed_list:
        label = p["filename"]
        if p.get("sheet_name"):
            label += f" [{p['sheet_name']}]"
        fname = p["filename"]
        ext = os.path.splitext(fname)[1] or ".xlsx"
        dl_name = f"{os.path.splitext(fname)[0]}_{p['sheet_name']}{ext}" if p.get("sheet_name") else fname
        st.download_button(
            label=f"📥 下载 {label}",
            data=p["_extracted_bytes"],
            file_name=dl_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"dl_{p['filename']}_{p.get('sheet_name', '')}",
        )

    # 显示审批标题（title 已在上方汇总表前生成）
    st.text_input("审批标题（自动生成）", value=title, disabled=True)

    # Submit button
    final_blocked = submit_blocked or signature_check_blocked
    if submit_blocked:
        st.error("⚠️ 校验未通过且当前为严格模式，请修改 Excel 后重新上传")
    if st.button("提交审批", disabled=final_blocked):
        progress_bar = st.progress(0)
        status_text = st.empty()

        try:
            open_id = st.session_state.open_id
            department_id = st.session_state.department_id

            file_codes = []
            # 总步数 = 各 sheet/附件 + 汇总表（如果生成成功） + 创建审批实例
            total_items = len(parsed_list)
            total_steps = total_items + (1 if summary_bytes else 0) + 1
            done_steps = 0

            for p in parsed_list:
                label = p["filename"]
                if p.get("sheet_name"):
                    label += f" [{p['sheet_name']}]"
                status_text.text(f"正在上传：{label} ...")

                sheet_bytes = p["_extracted_bytes"]

                # 取对应校验结果追加为验证 sheet
                val_key = _parsed_key(p)
                vr = validation_results.get(val_key)
                if (vr is not None
                    and val_cfg.get("enabled")
                    and val_cfg.get("write_back_sheet", True)):
                    sheet_bytes = append_validation_sheet(
                        sheet_bytes, vr,
                        sheet_name=val_cfg.get("write_back_sheet_name", "验证结果"),
                        source_filename=label,
                    )

                import tempfile
                fname = p["filename"]
                if p.get("sheet_name"):
                    fname = f"{os.path.splitext(fname)[0]}_{p['sheet_name']}.xlsx"
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                    tmp.write(sheet_bytes)
                    tmp_path = tmp.name
                file_code = client.upload_file_to_feishu(tmp_path, filename=fname)
                os.unlink(tmp_path)
                if file_code:
                    file_codes.append(file_code)
                done_steps += 1
                progress_bar.progress(done_steps / total_steps)

            # 把汇总表也作为额外附件上传（如果生成成功）
            if summary_bytes:
                status_text.text(f"正在上传：{summary_filename} ...")
                import tempfile
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                    tmp.write(summary_bytes)
                    tmp_path = tmp.name
                file_code = client.upload_file_to_feishu(tmp_path, filename=summary_filename)
                os.unlink(tmp_path)
                if file_code:
                    file_codes.append(file_code)
                done_steps += 1
                progress_bar.progress(done_steps / total_steps)

            status_text.text("正在创建审批实例...")
            form_json = client.build_feishu_form(parsed_list, title, file_codes, remark="")
            instance_code = client.create_approval_instance(
                open_id=open_id,
                department_id=department_id,
                form_json=form_json,
                approval_code=os.getenv("FEISHU_APPROVAL_CODE"),
            )
            progress_bar.progress(1.0)
            status_text.empty()

            if instance_code:
                st.success(f"审批创建成功！instanceCode：{instance_code}")
            else:
                st.warning("审批创建完成，但未返回 instanceCode")
        except Exception as e:
            progress_bar.empty()
            status_text.empty()
            st.error(f"提交失败：{e}")


def query_feishu_approval_definition():
    import lark_oapi as lark
    from lark_oapi.api.approval.v4 import GetApprovalRequest

    app_id = os.getenv("FEISHU_APP_ID")
    app_secret = os.getenv("FEISHU_APP_SECRET")
    approval_code = os.getenv("FEISHU_APPROVAL_CODE")

    if not app_id or not app_secret or not approval_code:
        print("Error: FEISHU_APP_ID, FEISHU_APP_SECRET, or FEISHU_APPROVAL_CODE not set")
        return None

    client = lark.Client.builder().app_id(app_id).app_secret(app_secret).build()

    request = GetApprovalRequest.builder() \
        .approval_code(approval_code) \
        .locale("zh-CN") \
        .user_id_type("open_id") \
        .build()

    response = client.approval.v4.approval.get(request)
    if not response.success():
        print(f"Error: code={response.code}, msg={response.msg}")
        return None

    form_str = response.data.form
    if not form_str:
        print("Error: empty form in response")
        return None

    try:
        controls = json.loads(form_str)
    except json.JSONDecodeError as e:
        print(f"Error parsing form JSON: {e}")
        return None

    _map_feishu_controls_to_config(controls)

    return [
        {"id": c.get("id"), "type": c.get("type"), "name": c.get("name")}
        for c in controls if isinstance(c, dict)
    ]


def _map_feishu_controls_to_config(controls):
    config_path = os.path.join(os.path.dirname(__file__), "config.json")
    if not os.path.exists(config_path):
        print("Warning: config.json not found, skipping mapping")
        return

    with open(config_path, "r", encoding="utf-8") as f:
        config = json.load(f)

    feishu_form = config.get("feishu_form", {})
    cfg_controls = feishu_form.get("controls", {})
    cfg_table_sub = feishu_form.get("table_sub_controls", {})

    if not cfg_controls and not cfg_table_sub:
        print("Warning: no feishu_form controls or table_sub_controls in config")
        return

    api_by_name = {}
    table_ctrl = None
    for ctrl in controls:
        if not isinstance(ctrl, dict):
            continue
        name = ctrl.get("name", "")
        ctrl_type = ctrl.get("type", "")
        if name:
            api_by_name[name] = ctrl
        if ctrl_type == "fieldList":
            table_ctrl = ctrl

    logger.debug("Available Feishu widget names: %s", list(api_by_name.keys()))
    if table_ctrl:
        logger.debug("fieldList raw keys: %s", list(table_ctrl.keys()))
        logger.debug("fieldList JSON: %s", json.dumps(table_ctrl, ensure_ascii=False)[:2000])

    unmatched = []

    # 如果配置中所有控件已有 feishu_id，跳过 API 查询
    all_mapped = all(cfg.get("feishu_id") for cfg in cfg_controls.values()) and \
                 all(cfg.get("feishu_id") for cfg in cfg_table_sub.values())
    if all_mapped:
        logger.info("All feishu controls already mapped, skipping API query")
        return

    for key, cfg in cfg_controls.items():
        hint = cfg.get("control_name_hint", "")
        if not hint or cfg.get("feishu_id"):
            continue
        matched = None
        if hint in api_by_name:
            matched = api_by_name[hint]
        else:
            for api_name, api_ctrl in api_by_name.items():
                if hint in api_name or api_name in hint:
                    matched = api_ctrl
                    break
        if matched:
            cfg["feishu_id"] = matched.get("id")
            print(f"Mapped control '{key}' -> '{matched.get('name')}' ({matched.get('id')})")
        else:
            unmatched.append((key, hint))

    if table_ctrl and cfg_table_sub:
        children = _extract_fieldlist_children(table_ctrl)
        if children:
            input_children = [c for c in children if isinstance(c, dict) and c.get("type") == "input"]
            amount_children = [c for c in children if isinstance(c, dict) and c.get("type") in ("number", "amount")]

            input_keys = [k for k, v in cfg_table_sub.items() if v.get("feishu_type") == "input"]
            for i, key in enumerate(input_keys):
                if i < len(input_children):
                    cfg_table_sub[key]["feishu_id"] = input_children[i].get("id")
                    print(
                        f"Mapped table_sub_control '{key}' -> "
                        f"'{input_children[i].get('name')}' ({input_children[i].get('id')})"
                    )
                else:
                    unmatched.append((key, f"input[{i}]"))

            amount_keys = [k for k, v in cfg_table_sub.items() if v.get("feishu_type") in ("number", "amount")]
            for i, key in enumerate(amount_keys):
                if i < len(amount_children):
                    cfg_table_sub[key]["feishu_id"] = amount_children[i].get("id")
                    print(
                        f"Mapped table_sub_control '{key}' -> "
                        f"'{amount_children[i].get('name')}' ({amount_children[i].get('id')})"
                    )
                else:
                    unmatched.append((key, f"amount[{i}]"))
        else:
            print("Warning: fieldList control found but no children extracted")
            for key in cfg_table_sub:
                unmatched.append((key, cfg_table_sub[key].get("feishu_type", "unknown")))

    if unmatched:
        print("Unmatched controls:")
        for key, hint in unmatched:
            print(f"  - {key}: {hint}")

    with open(config_path, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)
    print(f"Updated {config_path} with feishu_id mappings")


def _extract_fieldlist_children(table_ctrl):
    if not isinstance(table_ctrl, dict):
        return []

    ctrl_id = table_ctrl.get("id", "?")
    logger.debug("Extracting fieldList children for control id=%s", ctrl_id)

    # Try multiple known nesting paths
    candidates: list[dict] = []

    # Path 1: widget.children (most common)
    widget = table_ctrl.get("widget")
    if isinstance(widget, dict):
        for key in ("children", "sub_components", "subControls", "sub_controls"):
            children = widget.get(key)
            if isinstance(children, list):
                candidates = children
                logger.debug("  Found %d children via widget.%s", len(candidates), key)
                break

    # Path 2: direct children
    if not candidates:
        for key in ("children", "sub_components", "subControls", "sub_controls"):
            children = table_ctrl.get(key)
            if isinstance(children, list):
                candidates = children
                logger.debug("  Found %d children via %s", len(candidates), key)
                break

    # Path 3: recursive search through all nested dict values
    if not candidates:
        def _recurse_search(obj, depth=0):
            if depth > 5:
                return
            if isinstance(obj, dict):
                for k, v in obj.items():
                    if k in ("children", "sub_components", "subControls", "sub_controls"):
                        if isinstance(v, list) and any(isinstance(c, dict) and "id" in c for c in v):
                            candidates.extend(v)
                    _recurse_search(v, depth + 1)
            elif isinstance(obj, list):
                for item in obj:
                    _recurse_search(item, depth + 1)
        _recurse_search(table_ctrl)
        if candidates:
            logger.debug("  Found %d children via recursive search", len(candidates))

    if not candidates:
        logger.warning(
            "  No children found! Raw table_ctrl keys: %s",
            list(table_ctrl.keys()),
        )

    return candidates


if __name__ == "__main__":
    try:
        main()
    except Exception:
        logger.exception("main() 未捕获异常")
        raise
